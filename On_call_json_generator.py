import openpyxl
import datetime
import json
print("----------------------------------------------------------------------------------")
print("1. 24*7 Shift Pattern ")
print("2. General Shift Pattern [ for child sites | 8 am to 5 pm ]  ")
print("")
Site_Roaster_type = int(input("Enter your Roaster Type [Select 1 or 2 ] : "))
print("")
Excel_location = input("Enter your excel file location [Shortcut : option+cmd+C] :")

wb=openpyxl.load_workbook(Excel_location)
sheets = wb.sheetnames
Feb2023_sheet = wb['Sheet1']
row = Feb2023_sheet.max_row
col = Feb2023_sheet.max_column
Roaster_seq = []
exit_counter = 0
if Site_Roaster_type == 1 : 
    for i in range(2,row) :
        for j in range(3,6):
            on_call = Feb2023_sheet.cell(i,j).value
            if on_call != None:
                multi_on_call = on_call.split(",")
                Roaster_seq.append(multi_on_call)
            if on_call == None:
                exit_counter = 1
                exit
        if exit_counter ==1  :
            break

    Dates = []
    shift = 1
    roaster_i = 0

    Roaster_json = [ ]
    for i in range(2,row) :
        sample = {
            "startDateTime" : None,
            "endDateTime": None,
            "oncallMember": None
        }
        cur =  Feb2023_sheet.cell(i,1).value
        if cur != None or Roaster_seq[roaster_i] == None :
            start = cur.strftime("%m/%d/%Y") + " 7:00 "
            end = cur.strftime("%m/%d/%Y") + " 15:00 "
            sample["startDateTime"] = start
            sample["endDateTime"] = end
            sample["oncallMember"] = Roaster_seq[roaster_i]
            Roaster_json.append(sample)
            roaster_i = roaster_i+1
            # print(Roaster_json)
            sample = {
            "startDateTime" : None,
            "endDateTime": None,
            "oncallMember": None
            }
            
            start = cur.strftime("%m/%d/%Y") + " 15:00 "
            end = cur.strftime("%m/%d/%Y") + " 22:00"
            sample["startDateTime"] = start
            sample["endDateTime"] = end
            sample["oncallMember"] = Roaster_seq[roaster_i]
            Roaster_json.append(sample)
            roaster_i = roaster_i+1
            # print(Roaster_json)

            sample = {
            "startDateTime" : None,
            "endDateTime": None,
            "oncallMember": None
            }

            
            start = cur.strftime("%m/%d/%Y") + " 22:00 "
            cur_nxt = Feb2023_sheet.cell(i+1,1).value
            end = cur_nxt.strftime("%m/%d/%Y") + " 7:00"
            sample["startDateTime"] = start
            sample["endDateTime"] = end
            sample["oncallMember"] = Roaster_seq[roaster_i]
            Roaster_json.append(sample)
            roaster_i = roaster_i+1
            # print(Roaster_json)
            if Feb2023_sheet.cell(i+2,1).value == None :
                break
        else:
            break
    print("Congratulation!! Roaster made Successfully")
    with open("on_call_parent.json" , "w") as outfile :
        json.dump(Roaster_json, outfile)

if Site_Roaster_type == 2 :
    for i in range(2,row) :
        for j in range(3,5):
            on_call = Feb2023_sheet.cell(i,j).value
            if on_call != None:
                multi_on_call = on_call.split(",")
                Roaster_seq.append(multi_on_call)
            if on_call == None:
                exit_counter = 1
                exit
        if exit_counter ==1  :
            break

    Dates = []
    shift = 1
    roaster_i = 0

    Roaster_json = [ ]
    for i in range(2,row) :
        sample = {
            "startDateTime" : None,
            "endDateTime": None,
            "oncallMember": None
        }
        cur =  Feb2023_sheet.cell(i,1).value
        if cur != None or Roaster_seq[roaster_i] == None :
            start = cur.strftime("%m/%d/%Y") + " 8:00 "
            end = cur.strftime("%m/%d/%Y") + " 17:00 "
            sample["startDateTime"] = start
            sample["endDateTime"] = end
            sample["oncallMember"] = Roaster_seq[roaster_i]
            Roaster_json.append(sample)
            roaster_i = roaster_i+1
            

            sample = {
            "startDateTime" : None,
            "endDateTime": None,
            "oncallMember": None
            }

            
            start = cur.strftime("%m/%d/%Y") + " 17:00 "
            cur_nxt = Feb2023_sheet.cell(i+1,1).value
            end = cur_nxt.strftime("%m/%d/%Y") + " 8:00"
            sample["startDateTime"] = start
            sample["endDateTime"] = end
            sample["oncallMember"] = Roaster_seq[roaster_i]
            Roaster_json.append(sample)
            roaster_i = roaster_i+1
            # print(Roaster_json)
            if Feb2023_sheet.cell(i+2,1).value == None :
                break
        else:
            break
        date = datetime.datetime.today
        print

    print("Congratulation!! Roaster made Successfully")
    with open("on_call_child.json" , "w") as outfile :
        json.dump(Roaster_json, outfile) 